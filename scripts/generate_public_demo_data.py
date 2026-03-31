from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
WORKBOOK = ROOT / "Networking Tracker.xlsx"
OUTPUT = ROOT / "docs" / "data" / "contacts.json"


def mask_name(value: str) -> str:
    parts = [part for part in value.strip().split() if part]
    masked = []
    for part in parts:
        if len(part) <= 2:
            masked.append(part[0] + "*")
        else:
            masked.append(part[0] + "*" * (len(part) - 1))
    return " ".join(masked)


def build_contacts():
    workbook = load_workbook(WORKBOOK, data_only=True)
    contacts = []

    for sheet_name, school, tag in [
        ("Professional Networking", "Wharton", "professional-networking"),
        ("Casual Networking", "Penn", "casual-networking"),
    ]:
        sheet = workbook[sheet_name]
        headers = [sheet.cell(row=1, column=col).value for col in range(1, 14)]

        for row_idx in range(2, sheet.max_row + 1):
            values = [sheet.cell(row=row_idx, column=col).value for col in range(1, 14)]
            record = dict(zip(headers, values))
            name = (record.get("Name") or "").strip() if record.get("Name") else ""
            company = (record.get("Company") or "").strip() if record.get("Company") else ""
            position = (record.get("Position") or "").strip() if record.get("Position") else ""
            group = (record.get("Group") or "").strip() if record.get("Group") else ""

            if not name or not company:
                continue

            contacts.append(
                {
                    "id": f"{tag}-{len(contacts) + 1}",
                    "full_name": mask_name(name),
                    "firm": company or None,
                    "group": group or None,
                    "role": position or None,
                    "location": None,
                    "school": school,
                    "email": "Hidden for public demo" if record.get("Email") else "Unavailable",
                    "email_confidence_score": 0.95 if record.get("Email") else 0.2,
                    "linkedin_url": None,
                    "tags": [tag],
                }
            )

    return contacts


def main():
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "generated_from": WORKBOOK.name,
        "note": "Public demo dataset derived from the private networking tracker. Names and direct contact info are masked.",
        "contacts": build_contacts(),
    }
    OUTPUT.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote {len(payload['contacts'])} contacts to {OUTPUT}")


if __name__ == "__main__":
    main()
